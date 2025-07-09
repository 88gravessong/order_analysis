#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
使用 openpyxl 计算每个 SKU 的订单指标
------------------------------------------------
指标定义：
1. 订单数               = 该 SKU 的所有订单数
2. 已完成率(%)          = Order Substatus == "已完成" 且 Cancelation/Return Type 为空 的订单数   / 订单数 * 100
3. 已送达率(%)          = Order Substatus == "已送达"                                   / 订单数 * 100
4. 退款率(%)            = Order Substatus 含 "Return"/"Refund"                          / 订单数 * 100
5. 发货前取消率(%)      = Order Substatus == "已取消" 且 Shipped Time 为空              / 订单数 * 100
6. 发货后取消率(%)      = Order Substatus == "已取消" 且 Shipped Time 不为空            / 订单数 * 100
7. 仍在途率(%)          = Order Substatus == "运输中"                                   / 订单数 * 100
8. 签收率(%)            = 已完成率 + 已送达率 + 退款率

注意：
- 表格第二行是描述行，需要跳过。
- 仅依赖 openpyxl 进行 Excel 读写，避免只读取到第一列的问题。
"""

from collections import defaultdict
from pathlib import Path
from typing import Dict, Sequence

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

INPUT_FILE = "全部 订单-2025-07-08-21_50.xlsx"  # 如需处理其它文件，可修改此常量或传参
OUTPUT_FILE = "订单指标分析结果.xlsx"

# 需要用到的列名（不区分大小写）
TARGET_COLUMNS = {
    "order_substatus": ["order substatus"],
    "cancel_type": ["cancelation/return type", "cancellation/return type"],
    "seller_sku": ["seller sku"],  # 以 Seller SKU 为分组键
    "shipped_time": ["shipped time"],
}


def normalise(text: object) -> str:
    """统一大小写并去除多余空白

    参数接收任意对象以防传入 None/非字符串。返回经 strip+lower 处理后的字符串。"""
    if isinstance(text, str):
        return text.strip().lower()
    return ""


def locate_columns(headers: Sequence[str]) -> Dict[str, int]:
    """根据标题行定位目标列索引 (0-based)"""
    header_map = {normalise(h): idx for idx, h in enumerate(headers) if h}

    col_idx_map: Dict[str, int] = {}
    for key, aliases in TARGET_COLUMNS.items():
        for alias in aliases:
            if alias in header_map:
                col_idx_map[key] = header_map[alias]
                break
        if key not in col_idx_map:
            raise KeyError(f"未找到列: {aliases[0]} (实际标题行: {headers})")
    return col_idx_map


def read_orders(file_path: Path):
    """读取 Excel，返回迭代器 (sku_id, substatus, cancel_type, shipped_time)"""
    wb = load_workbook(file_path, data_only=True)
    ws: Worksheet = wb.active  # type: ignore[assignment]

    # 正确读取标题行（read_only 模式下无法通过 ws[1] 获取完整行）
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header_row]
    col_indices = locate_columns(headers)

    def _safe(row, idx):
        return row[idx] if idx < len(row) else None

    # 从第3行开始遍历（第2行是描述行）
    for row in ws.iter_rows(min_row=3, values_only=True):  # type: ignore[attr-defined]
        seller_sku = _safe(row, col_indices["seller_sku"])
        substatus = _safe(row, col_indices["order_substatus"])
        cancel_type = _safe(row, col_indices["cancel_type"])
        shipped_time = _safe(row, col_indices["shipped_time"])
        yield seller_sku, substatus, cancel_type, shipped_time

    wb.close()


def compute_metrics(file_path: Path):
    """核心计算逻辑"""
    stats: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))

    total_rows = 0
    for seller_sku, sub, cancel, shipped in read_orders(file_path):
        if seller_sku is None:
            continue  # 跳过无效行
        sku_id = str(seller_sku)
        total_rows += 1
        s = stats[sku_id]
        s["total"] += 1

        sub = normalise(sub)
        cancel = normalise(cancel)
        shipped_empty = shipped is None or str(shipped).strip() == ""

        if sub == "已完成" and cancel == "":
            s["completed"] += 1
        elif sub == "已送达":
            s["delivered"] += 1
        elif "return" in sub or "refund" in sub:
            s["refund"] += 1
        elif sub == "已取消":
            if shipped_empty:
                s["cancel_before"] += 1
            else:
                s["cancel_after"] += 1
        elif sub == "运输中":
            s["in_transit"] += 1

        # 其它状态直接忽略

    print(f"已读取 {total_rows} 行订单记录，发现 {len(stats)} 个 SKU")
    return stats


def build_result_workbook(stats: Dict[str, Dict[str, int]]) -> Workbook:
    """根据统计结果构建结果工作簿"""
    wb = Workbook()

    # openpyxl 保证 .active 返回 Worksheet，但为稳妥起见加断言并显式标注类型
    ws_raw = wb.active
    assert ws_raw is not None, "无法获取默认工作表"
    ws: Worksheet = ws_raw  # type: ignore[assignment]

    # 安全设置工作表标题
    try:
        ws.title = "订单指标"
    except Exception:
        # 如果设置失败，保持默认标题
        pass

    headers = [
        "Seller SKU",
        "订单数",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]
    try:
        ws.append(headers)
    except ValueError:
        # 在极罕见情况下 openpyxl 会因为空表或维度问题抛 ValueError
        pass

    # 写数据
    for sku, m in sorted(stats.items(), key=lambda x: (-x[1]["total"], x[0])):
        total = m["total"]
        completed_rate = m["completed"] / total * 100 if total else 0
        delivered_rate = m["delivered"] / total * 100 if total else 0
        refund_rate = m["refund"] / total * 100 if total else 0
        cancel_before_rate = m["cancel_before"] / total * 100 if total else 0
        cancel_after_rate = m["cancel_after"] / total * 100 if total else 0
        in_transit_rate = m["in_transit"] / total * 100 if total else 0
        sign_rate = completed_rate + delivered_rate + refund_rate

        try:
            ws.append([
                sku,
                total,
                round(sign_rate, 2),
                round(completed_rate, 2),
                round(delivered_rate, 2),
                round(refund_rate, 2),
                round(cancel_before_rate, 2),
                round(cancel_after_rate, 2),
                round(in_transit_rate, 2),
            ])
        except ValueError:
            # 忽略无法写入的行
            continue

    # 自动调整列宽
    for col_idx, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 14

    return wb


def main():
    print("=== 订单指标计算程序 (openpyxl) ===")
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise FileNotFoundError(f"找不到输入文件: {INPUT_FILE}")

    stats = compute_metrics(input_path)
    if not stats:
        print("未找到任何可计算数据。")
        return

    wb = build_result_workbook(stats)
    wb.save(OUTPUT_FILE)
    print(f"计算完成，结果已保存为: {OUTPUT_FILE}")


if __name__ == "__main__":
    main() 